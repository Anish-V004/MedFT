"""
PV Model Evaluation Script (Manual / Rule-Based)
==================================================
Parses the JSONL dataset, extracts structured JSON from both
ground-truth (assistant) and fine-tuned model responses, compares
them field-by-field, and writes results to an Excel workbook with:
  • Sheet 1 – "Comparison Logs": row-by-row field comparison
  • Sheet 2 – "Analytics Summary": overall accuracy metrics
"""

import json
import re
import os
import sys
from collections import Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────
DATASET  = os.path.join(os.path.dirname(__file__), "pv_test_results_300.jsonl")
OUTPUT   = os.path.join(os.path.dirname(__file__), "pv_evaluation_results.xlsx")

# ── Colour palette ─────────────────────────────────────────────────
GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ── Helpers ────────────────────────────────────────────────────────

def extract_json_block(text: str) -> dict | None:
    """Pull the first ```json ... ``` block and parse it."""
    m = re.search(r"```json\s*(\{.*?\})\s*```", text, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError:
            return None
    return None


def normalise(val):
    """Lower-case, strip, collapse whitespace for comparison."""
    if val is None:
        return ""
    return re.sub(r"\s+", " ", str(val).strip().lower())


def compare_seriousness(gt: dict, pred: dict) -> tuple[bool, bool, str]:
    """
    Returns (is_serious_match, criteria_match, detail_string).
    """
    gt_s = gt.get("seriousness", {})
    pr_s = pred.get("seriousness", {})

    gt_serious  = gt_s.get("is_serious")
    pr_serious  = pr_s.get("is_serious")
    serious_match = (gt_serious == pr_serious)

    gt_crit = normalise(gt_s.get("criteria", ""))
    pr_crit = normalise(pr_s.get("criteria", ""))
    criteria_match = (gt_crit == pr_crit)

    detail = f"GT: serious={gt_serious}, criteria='{gt_s.get('criteria','')}' | "
    detail += f"Pred: serious={pr_serious}, criteria='{pr_s.get('criteria','')}'"
    return serious_match, criteria_match, detail


def compare_meddra(gt: dict, pred: dict) -> tuple[bool, str]:
    gt_pt = normalise(gt.get("meddra_pt", ""))
    pr_pt = normalise(pred.get("meddra_pt", ""))
    match = (gt_pt == pr_pt)
    detail = f"GT: '{gt.get('meddra_pt','')}' | Pred: '{pred.get('meddra_pt','')}'"
    return match, detail


def compare_expectedness(gt: dict, pred: dict) -> tuple[bool, str]:
    gt_e = normalise(gt.get("expectedness", ""))
    pr_e = normalise(pred.get("expectedness", ""))
    match = (gt_e == pr_e)
    detail = f"GT: '{gt.get('expectedness','')}' | Pred: '{pred.get('expectedness','')}'"
    return match, detail


def compare_causality(gt: dict, pred: dict) -> tuple[bool, bool, str]:
    """
    Returns (naranjo_score_match, interpretation_match, detail_string).
    """
    gt_c = gt.get("causality", {})
    pr_c = pred.get("causality", {})

    gt_score = gt_c.get("naranjo_score")
    pr_score = pr_c.get("naranjo_score")
    score_match = (gt_score == pr_score)

    gt_interp = normalise(gt_c.get("interpretation", ""))
    pr_interp = normalise(pr_c.get("interpretation", ""))
    interp_match = (gt_interp == pr_interp)

    detail  = f"GT: score={gt_score}, interp='{gt_c.get('interpretation','')}' | "
    detail += f"Pred: score={pr_score}, interp='{pr_c.get('interpretation','')}'"
    return score_match, interp_match, detail


def naranjo_bucket(score) -> str:
    """Map a Naranjo score to its interpretation category."""
    if score is None:
        return "N/A"
    try:
        s = int(score)
    except (ValueError, TypeError):
        return "N/A"
    if s >= 9:
        return "Definite"
    if s >= 5:
        return "Probable"
    if s >= 1:
        return "Possible"
    return "Doubtful"


def score_diff(gt_score, pr_score) -> int | str:
    """Absolute difference between Naranjo scores."""
    try:
        return abs(int(gt_score) - int(pr_score))
    except (ValueError, TypeError):
        return "N/A"

# ── Main ───────────────────────────────────────────────────────────

def main():
    # 1. Load data
    rows = []
    with open(DATASET, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            rows.append(json.loads(line))

    print(f"Loaded {len(rows)} rows from dataset.")

    # 2. Parse & compare
    results = []
    parse_failures_gt = 0
    parse_failures_pred = 0

    for idx, row in enumerate(rows, start=1):
        msgs = row.get("messages", [])

        # Extract fields
        user_msg = ""
        gt_text  = ""
        pred_text = ""
        for m in msgs:
            if m["role"] == "user":
                user_msg = m["content"]
            elif m["role"] == "assistant":
                gt_text = m["content"]
            elif m["role"] == "trained_model_response":
                pred_text = m["content"]

        # Extract patient narrative snippet (first 200 chars)
        narr_match = re.search(r"Patient Narrative:\s*(.+?)(?:\n\nReference Safety|$)", user_msg, re.DOTALL)
        narrative_snippet = narr_match.group(1)[:200].strip() if narr_match else user_msg[:200]

        # Extract suspected drug
        drug_match = re.search(r"The suspected drug is ([^.]+)\.", user_msg)
        suspected_drug = drug_match.group(1).strip() if drug_match else "N/A"

        # Parse JSON blocks
        gt_json = extract_json_block(gt_text)
        pr_json = extract_json_block(pred_text)

        if gt_json is None:
            parse_failures_gt += 1
        if pr_json is None:
            parse_failures_pred += 1

        if gt_json is None or pr_json is None:
            results.append({
                "row": idx,
                "drug": suspected_drug,
                "narrative": narrative_snippet,
                "parse_error": True,
                "gt_json": gt_json,
                "pr_json": pr_json,
            })
            continue

        # Compare
        serious_match, criteria_match, serious_detail = compare_seriousness(gt_json, pr_json)
        meddra_match, meddra_detail = compare_meddra(gt_json, pr_json)
        expect_match, expect_detail = compare_expectedness(gt_json, pr_json)
        naranjo_match, interp_match, causal_detail = compare_causality(gt_json, pr_json)

        # Naranjo bucket match (interpretation category)
        gt_naranjo = gt_json.get("causality", {}).get("naranjo_score")
        pr_naranjo = pr_json.get("causality", {}).get("naranjo_score")
        gt_bucket = naranjo_bucket(gt_naranjo)
        pr_bucket = naranjo_bucket(pr_naranjo)
        bucket_match = (normalise(gt_bucket) == normalise(pr_bucket))

        # Overall match (all critical fields)
        overall = all([serious_match, criteria_match, meddra_match, expect_match, interp_match])

        results.append({
            "row": idx,
            "drug": suspected_drug,
            "narrative": narrative_snippet,
            "parse_error": False,
            # Seriousness
            "gt_is_serious": gt_json.get("seriousness", {}).get("is_serious"),
            "pr_is_serious": pr_json.get("seriousness", {}).get("is_serious"),
            "serious_match": serious_match,
            "gt_criteria": gt_json.get("seriousness", {}).get("criteria", ""),
            "pr_criteria": pr_json.get("seriousness", {}).get("criteria", ""),
            "criteria_match": criteria_match,
            # MedDRA
            "gt_meddra": gt_json.get("meddra_pt", ""),
            "pr_meddra": pr_json.get("meddra_pt", ""),
            "meddra_match": meddra_match,
            # Expectedness
            "gt_expect": gt_json.get("expectedness", ""),
            "pr_expect": pr_json.get("expectedness", ""),
            "expect_match": expect_match,
            # Causality
            "gt_naranjo": gt_naranjo,
            "pr_naranjo": pr_naranjo,
            "naranjo_match": naranjo_match,
            "naranjo_diff": score_diff(gt_naranjo, pr_naranjo),
            "gt_interp": gt_json.get("causality", {}).get("interpretation", ""),
            "pr_interp": pr_json.get("causality", {}).get("interpretation", ""),
            "interp_match": interp_match,
            "gt_bucket": gt_bucket,
            "pr_bucket": pr_bucket,
            "bucket_match": bucket_match,
            # Overall
            "overall_match": overall,
        })

    # ── 3. Write Excel ─────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ── Sheet 1: Comparison Logs ───────────────────────────────────
    ws1 = wb.active
    ws1.title = "Comparison Logs"

    headers = [
        "Row #", "Suspected Drug", "Narrative (Snippet)",
        "GT is_serious", "Pred is_serious", "Seriousness Match",
        "GT Criteria", "Pred Criteria", "Criteria Match",
        "GT MedDRA PT", "Pred MedDRA PT", "MedDRA Match",
        "GT Expectedness", "Pred Expectedness", "Expectedness Match",
        "GT Naranjo Score", "Pred Naranjo Score", "Naranjo Score Match", "Naranjo Δ",
        "GT Interpretation", "Pred Interpretation", "Interpretation Match",
        "GT Naranjo Bucket", "Pred Naranjo Bucket", "Bucket Match",
        "Overall Match", "Parse Error",
    ]

    for col_idx, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    for r_idx, res in enumerate(results, start=2):
        if res["parse_error"]:
            vals = [
                res["row"], res["drug"], res["narrative"],
                "", "", "", "", "", "", "", "", "", "", "", "",
                "", "", "", "", "", "", "", "", "", "",
                "", True,
            ]
        else:
            vals = [
                res["row"], res["drug"], res["narrative"],
                res["gt_is_serious"], res["pr_is_serious"], res["serious_match"],
                res["gt_criteria"], res["pr_criteria"], res["criteria_match"],
                res["gt_meddra"], res["pr_meddra"], res["meddra_match"],
                res["gt_expect"], res["pr_expect"], res["expect_match"],
                res["gt_naranjo"], res["pr_naranjo"], res["naranjo_match"], res["naranjo_diff"],
                res["gt_interp"], res["pr_interp"], res["interp_match"],
                res["gt_bucket"], res["pr_bucket"], res["bucket_match"],
                res["overall_match"], res["parse_error"],
            ]
        for c_idx, v in enumerate(vals, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            # Colour match columns
            if isinstance(v, bool) and c_idx not in [4, 5]:  # skip is_serious booleans
                cell.fill = GREEN if v else RED

    # Auto-width (capped)
    for col_idx in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = min(
            max(len(str(headers[col_idx - 1])) + 4, 14), 40
        )
    ws1.column_dimensions["C"].width = 60  # narrative column wider
    ws1.auto_filter.ref = ws1.dimensions
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Analytics Summary ─────────────────────────────────
    ws2 = wb.create_sheet("Analytics Summary")

    valid = [r for r in results if not r["parse_error"]]
    total = len(results)
    valid_count = len(valid)
    parse_err_count = total - valid_count

    def acc(key):
        if not valid:
            return 0
        return sum(1 for r in valid if r.get(key)) / valid_count

    def count_match(key):
        return sum(1 for r in valid if r.get(key))

    # --- Section: High-Level Summary ---
    section_row = 1
    ws2.cell(row=section_row, column=1, value="PV MODEL EVALUATION – ANALYTICS SUMMARY").font = Font(bold=True, size=14)
    ws2.merge_cells(start_row=section_row, start_column=1, end_row=section_row, end_column=4)

    section_row = 3
    summary_headers = ["Metric", "Value", "Percentage"]
    for ci, h in enumerate(summary_headers, 1):
        c = ws2.cell(row=section_row, column=ci, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = THIN_BORDER

    summary_data = [
        ("Total Samples", total, ""),
        ("Valid (Parsed OK)", valid_count, f"{valid_count/total*100:.1f}%" if total else ""),
        ("Parse Failures", parse_err_count, f"{parse_err_count/total*100:.1f}%" if total else ""),
        ("", "", ""),
        ("── FIELD-LEVEL ACCURACY ──", "", ""),
        ("Seriousness (is_serious)", count_match("serious_match"), f"{acc('serious_match')*100:.1f}%"),
        ("Seriousness Criteria", count_match("criteria_match"), f"{acc('criteria_match')*100:.1f}%"),
        ("MedDRA PT (Exact Match)", count_match("meddra_match"), f"{acc('meddra_match')*100:.1f}%"),
        ("Expectedness", count_match("expect_match"), f"{acc('expect_match')*100:.1f}%"),
        ("Naranjo Score (Exact)", count_match("naranjo_match"), f"{acc('naranjo_match')*100:.1f}%"),
        ("Naranjo Interpretation", count_match("interp_match"), f"{acc('interp_match')*100:.1f}%"),
        ("Naranjo Bucket Match", count_match("bucket_match"), f"{acc('bucket_match')*100:.1f}%"),
        ("", "", ""),
        ("── COMPOSITE ──", "", ""),
        ("All Fields Exact Match", count_match("overall_match"), f"{acc('overall_match')*100:.1f}%"),
    ]

    for ri, (metric, val, pct) in enumerate(summary_data, start=section_row + 1):
        ws2.cell(row=ri, column=1, value=metric).font = BOLD if "──" in str(metric) else Font(size=11)
        ws2.cell(row=ri, column=2, value=val).border = THIN_BORDER
        ws2.cell(row=ri, column=3, value=pct).border = THIN_BORDER
        ws2.cell(row=ri, column=1).border = THIN_BORDER
        # Highlight accuracy rows
        if isinstance(pct, str) and "%" in pct:
            pct_val = float(pct.replace("%", ""))
            fill = GREEN if pct_val >= 80 else (YELLOW if pct_val >= 60 else RED)
            ws2.cell(row=ri, column=3).fill = fill

    # --- Section: Naranjo Score Distribution ---
    naranjo_section_start = section_row + len(summary_data) + 3
    ws2.cell(row=naranjo_section_start, column=1, value="NARANJO SCORE DEVIATION DISTRIBUTION").font = Font(bold=True, size=12)

    dev_headers = ["Δ (Absolute Diff)", "Count", "% of Valid"]
    for ci, h in enumerate(dev_headers, 1):
        c = ws2.cell(row=naranjo_section_start + 1, column=ci, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = THIN_BORDER

    diffs = [r["naranjo_diff"] for r in valid if isinstance(r.get("naranjo_diff"), int)]
    diff_counter = Counter(diffs)
    for ri, diff_val in enumerate(sorted(diff_counter.keys()), start=naranjo_section_start + 2):
        cnt = diff_counter[diff_val]
        ws2.cell(row=ri, column=1, value=diff_val).border = THIN_BORDER
        ws2.cell(row=ri, column=2, value=cnt).border = THIN_BORDER
        ws2.cell(row=ri, column=3, value=f"{cnt/valid_count*100:.1f}%").border = THIN_BORDER

    # --- Section: Expectedness Confusion ---
    exp_section_start = naranjo_section_start + len(diff_counter) + 4
    ws2.cell(row=exp_section_start, column=1, value="EXPECTEDNESS CONFUSION MATRIX").font = Font(bold=True, size=12)

    exp_labels = sorted(set(
        [normalise(r["gt_expect"]) for r in valid] + [normalise(r["pr_expect"]) for r in valid]
    ))
    # Header row
    ws2.cell(row=exp_section_start + 1, column=1, value="GT \\ Pred").fill = HEADER_FILL
    ws2.cell(row=exp_section_start + 1, column=1).font = HEADER_FONT
    ws2.cell(row=exp_section_start + 1, column=1).border = THIN_BORDER
    for ci, lbl in enumerate(exp_labels, 2):
        c = ws2.cell(row=exp_section_start + 1, column=ci, value=lbl)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = THIN_BORDER

    for ri, gt_lbl in enumerate(exp_labels, start=exp_section_start + 2):
        ws2.cell(row=ri, column=1, value=gt_lbl).font = BOLD
        ws2.cell(row=ri, column=1).border = THIN_BORDER
        for ci, pr_lbl in enumerate(exp_labels, 2):
            cnt = sum(1 for r in valid if normalise(r["gt_expect"]) == gt_lbl and normalise(r["pr_expect"]) == pr_lbl)
            cell = ws2.cell(row=ri, column=ci, value=cnt)
            cell.border = THIN_BORDER
            if gt_lbl == pr_lbl:
                cell.fill = GREEN

    # --- Section: Causality Interpretation Confusion ---
    caus_section_start = exp_section_start + len(exp_labels) + 4
    ws2.cell(row=caus_section_start, column=1, value="CAUSALITY INTERPRETATION CONFUSION MATRIX").font = Font(bold=True, size=12)

    interp_labels = sorted(set(
        [normalise(r["gt_interp"]) for r in valid] + [normalise(r["pr_interp"]) for r in valid]
    ))
    ws2.cell(row=caus_section_start + 1, column=1, value="GT \\ Pred").fill = HEADER_FILL
    ws2.cell(row=caus_section_start + 1, column=1).font = HEADER_FONT
    ws2.cell(row=caus_section_start + 1, column=1).border = THIN_BORDER
    for ci, lbl in enumerate(interp_labels, 2):
        c = ws2.cell(row=caus_section_start + 1, column=ci, value=lbl)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = THIN_BORDER

    for ri, gt_lbl in enumerate(interp_labels, start=caus_section_start + 2):
        ws2.cell(row=ri, column=1, value=gt_lbl).font = BOLD
        ws2.cell(row=ri, column=1).border = THIN_BORDER
        for ci, pr_lbl in enumerate(interp_labels, 2):
            cnt = sum(1 for r in valid if normalise(r["gt_interp"]) == gt_lbl and normalise(r["pr_interp"]) == pr_lbl)
            cell = ws2.cell(row=ri, column=ci, value=cnt)
            cell.border = THIN_BORDER
            if gt_lbl == pr_lbl:
                cell.fill = GREEN

    # --- Section: Seriousness Criteria Distribution ---
    crit_section_start = caus_section_start + len(interp_labels) + 4
    ws2.cell(row=crit_section_start, column=1, value="SERIOUSNESS CRITERIA DISTRIBUTION").font = Font(bold=True, size=12)

    crit_headers = ["Criteria (GT)", "Count", "Model Matched"]
    for ci, h in enumerate(crit_headers, 1):
        c = ws2.cell(row=crit_section_start + 1, column=ci, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = THIN_BORDER

    gt_criteria_counts = Counter(r["gt_criteria"] for r in valid)
    for ri, (crit, cnt) in enumerate(gt_criteria_counts.most_common(), start=crit_section_start + 2):
        matched = sum(1 for r in valid if r["gt_criteria"] == crit and r["criteria_match"])
        ws2.cell(row=ri, column=1, value=crit).border = THIN_BORDER
        ws2.cell(row=ri, column=2, value=cnt).border = THIN_BORDER
        ws2.cell(row=ri, column=3, value=f"{matched}/{cnt} ({matched/cnt*100:.0f}%)").border = THIN_BORDER

    # --- Section: Mean Naranjo Deviation ---
    mean_section_start = crit_section_start + len(gt_criteria_counts) + 4
    ws2.cell(row=mean_section_start, column=1, value="NARANJO SCORE STATISTICS").font = Font(bold=True, size=12)

    numeric_diffs = [r["naranjo_diff"] for r in valid if isinstance(r.get("naranjo_diff"), int)]
    if numeric_diffs:
        mean_diff = sum(numeric_diffs) / len(numeric_diffs)
        max_diff = max(numeric_diffs)
        median_diff = sorted(numeric_diffs)[len(numeric_diffs) // 2]
    else:
        mean_diff = max_diff = median_diff = 0

    stats = [
        ("Mean Absolute Deviation", f"{mean_diff:.2f}"),
        ("Median Absolute Deviation", str(median_diff)),
        ("Max Absolute Deviation", str(max_diff)),
        ("Exact Score Match Rate", f"{acc('naranjo_match')*100:.1f}%"),
        ("Within ±1 Match Rate", f"{sum(1 for d in numeric_diffs if d <= 1)/len(numeric_diffs)*100:.1f}%" if numeric_diffs else "N/A"),
        ("Within ±2 Match Rate", f"{sum(1 for d in numeric_diffs if d <= 2)/len(numeric_diffs)*100:.1f}%" if numeric_diffs else "N/A"),
    ]

    for ri, (lbl, val) in enumerate(stats, start=mean_section_start + 1):
        ws2.cell(row=ri, column=1, value=lbl).font = BOLD
        ws2.cell(row=ri, column=1).border = THIN_BORDER
        ws2.cell(row=ri, column=2, value=val).border = THIN_BORDER

    # Column widths for analytics sheet
    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 18
    ws2.column_dimensions["E"].width = 18

    # ── Save ───────────────────────────────────────────────────────
    wb.save(OUTPUT)
    print(f"\n{'='*60}")
    print(f"  Evaluation complete!")
    print(f"  Total samples:       {total}")
    print(f"  Valid (parsed OK):   {valid_count}")
    print(f"  Parse failures:      {parse_err_count}")
    print(f"{'='*60}")
    print(f"  FIELD ACCURACIES:")
    print(f"    Seriousness:       {acc('serious_match')*100:.1f}%")
    print(f"    Criteria:          {acc('criteria_match')*100:.1f}%")
    print(f"    MedDRA PT:         {acc('meddra_match')*100:.1f}%")
    print(f"    Expectedness:      {acc('expect_match')*100:.1f}%")
    print(f"    Naranjo Score:     {acc('naranjo_match')*100:.1f}%")
    print(f"    Naranjo Interp:    {acc('interp_match')*100:.1f}%")
    print(f"    Overall (all):     {acc('overall_match')*100:.1f}%")
    print(f"{'='*60}")
    print(f"  Results saved to: {OUTPUT}")


if __name__ == "__main__":
    main()
