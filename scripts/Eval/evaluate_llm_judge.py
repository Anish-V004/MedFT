"""
PV Model Evaluation – LLM-as-Judge (gemini-3.1-flash-lite)
===========================================================
Uses Gemini as a clinical judge to evaluate the fine-tuned model's
responses against ground truth using clinical correctness (1-5 scale).

Key-rotation logic mirrors generate_reviews.py:
  - Round-robin queue across all keys
  - Per-key consecutive failure tracking; blacklist after 5 failures
  - If ALL active keys hit rate limits consecutively, sleep 65s (up to 3x)
  - After 3 sleeps → fatal exit
"""

import json
import re
import os
import sys
import time
import queue
import argparse
from collections import Counter
from dotenv import load_dotenv
from pydantic import BaseModel, Field

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from google import genai
from google.genai import types

# Reconfigure stdout to UTF-8 to prevent console encoding errors on Windows
sys.stdout.reconfigure(encoding="utf-8")

# Load environment variables
load_dotenv()

# ── Paths ──────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(__file__)
DATASET    = os.path.join(BASE_DIR, "pv_test_results_300.jsonl")
OUTPUT     = os.path.join(BASE_DIR, "pv_evaluation_llm_judge.xlsx")
CHECKPOINT = os.path.join(BASE_DIR, "llm_judge_checkpoint.json")

MODEL_NAME = "gemini-3.1-flash-lite"

# ── Styles ─────────────────────────────────────────────────────────
GREEN       = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED         = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW      = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD        = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ── Pydantic Schema for Structured LLM Judge Response ──────────────
class LLMJudgeResponse(BaseModel):
    seriousness_is_serious_match:   bool = Field(description="True if the model's is_serious is clinically correct or close enough based on context.")
    seriousness_criteria_match:     bool = Field(description="True if seriousness criteria is clinically correct or close enough.")
    meddra_pt_match:                bool = Field(description="True if MedDRA PT is clinically correct or close enough.")
    expectedness_match:             bool = Field(description="True if expectedness is clinically correct or close enough.")
    naranjo_score_match:            bool = Field(description="True if Naranjo score is clinically correct or close enough (within +/- 1 and clinically justified).")
    naranjo_interpretation_match:   bool = Field(description="True if Naranjo interpretation is correct or close enough.")
    clinical_accuracy_score:        int  = Field(description="Clinical accuracy score on a scale from 1 (unacceptable) to 5 (excellent/perfect).")
    clinical_explanation:           str  = Field(description="Detailed explanation of why this score was assigned, including clinical justification and strengths/weaknesses.")


# ── API key loading (mirrors generate_reviews.py) ──────────────────
def load_api_keys() -> list[str]:
    keys_str = os.environ.get("GEMINI_API_KEYS") or os.environ.get("GEMINI_API_KEY")
    keys: list[str] = []
    if keys_str:
        keys = [k.strip() for k in re.split(r"[,;]", keys_str) if k.strip()]
    for idx in range(1, 31):
        k = os.environ.get(f"GEMINI_API_KEY_{idx}")
        if k and k.strip() and k.strip() not in keys:
            keys.append(k.strip())
    return keys


# ── Key-rotation state (same pattern as generate_reviews.py) ───────
api_keys: list[str] = []
key_names: dict[str, str] = {}
client_queue: queue.Queue = queue.Queue()
key_last_used: dict[str, float] = {}
key_consecutive_failures: dict[str, int] = {}
disabled_keys: set[str] = set()
consecutive_failures = 0
consecutive_sleeps   = 0
MIN_INTERVAL = 4.0  # seconds between requests per key


def init_key_manager(keys: list[str]):
    global api_keys, key_names, client_queue, key_last_used
    global key_consecutive_failures, disabled_keys, consecutive_failures, consecutive_sleeps
    api_keys = keys
    key_names = {k: f"Key {i+1}" for i, k in enumerate(keys)}
    client_queue = queue.Queue()
    for k in keys:
        client_queue.put((k, genai.Client(api_key=k)))
    key_last_used = {k: 0.0 for k in keys}
    key_consecutive_failures = {k: 0 for k in keys}
    disabled_keys = set()
    consecutive_failures = 0
    consecutive_sleeps   = 0
    print(f"Loaded {len(keys)} API key(s): {[key_names[k] for k in keys]}")


def handle_key_failure(key_val: str, error_msg: str):
    global disabled_keys
    key_consecutive_failures[key_val] += 1
    failures = key_consecutive_failures[key_val]
    print(f"  [Key Failure] {key_names[key_val]} failed. Consecutive failures: {failures}/5.")
    sys.stdout.flush()
    if failures >= 5:
        disabled_keys.add(key_val)
        print(f"  [Key Disabled] {key_names[key_val]} has 5 consecutive failures – removing from rotation.")
        sys.stdout.flush()
        if len(disabled_keys) == len(api_keys):
            print(f"\n[FATAL] All {len(api_keys)} API keys disabled. Terminating.")
            sys.stdout.flush()
            os._exit(1)


def handle_rate_limit(key_val: str):
    global consecutive_failures, consecutive_sleeps
    consecutive_failures += 1
    active_count = len(api_keys) - len(disabled_keys)
    if consecutive_failures >= active_count:
        consecutive_sleeps += 1
        if consecutive_sleeps >= 3:
            print("\n[FATAL] All keys exhausted quota after 3 sleep cycles. Terminating.")
            sys.stdout.flush()
            os._exit(1)
        print(f"\nAll active keys hit rate-limit consecutively (sleep #{consecutive_sleeps}/3). Sleeping 65s...")
        sys.stdout.flush()
        time.sleep(65)
        consecutive_failures = 0
    else:
        print(f"  [Rate Limit] {key_names[key_val]} – rotating... (attempt {consecutive_failures}/{active_count})")
        sys.stdout.flush()
        time.sleep(1)


def handle_success(key_val: str):
    global consecutive_failures, consecutive_sleeps
    consecutive_failures = 0
    consecutive_sleeps   = 0
    key_consecutive_failures[key_val] = 0


# ── JSON extraction ────────────────────────────────────────────────
def extract_json_block(text: str) -> dict | None:
    m = re.search(r"```json\s*(\{.*?\})\s*```", text, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError:
            return None
    return None


# ── Judge system prompt ────────────────────────────────────────────
JUDGE_SYSTEM = """You are an expert Pharmacovigilance (PV) evaluator and clinical safety judge. You will be given:
1. A Patient Narrative and RSI (context)
2. A Ground Truth JSON assessment (created by an expert)
3. A Model Predicted response (which contains clinical Chain of Thought reasoning + JSON block)

Your job is to read the patient narrative, RSI, Ground Truth, and Model response. Then, use your clinical knowledge and standard PV rules to evaluate if the Model's findings and clinical reasoning are correct, logically sound, and clinically close enough to the Ground Truth.

Guidelines for "Correct or Clinically Close Enough":
- MedDRA PT: Accept standard clinical synonyms or minor wording variations if they represent the same clinical event (e.g., "Cerebral haemorrhage" is close enough and clinically equivalent to "Intracranial haemorrhage" or "Cerebellar hemorrhage").
- Seriousness is_serious: Must be clinically correct based on standard regulatory criteria.
- Seriousness criteria: Accept semantically equivalent terms (e.g., "hospitalization" and "hospitalisation"; "other serious medical event" and "other medically important condition").
- Expectedness: Must be clinically correct based on strictly comparing the event against the provided RSI.
- Naranjo score: Allow a tolerance of +/- 1 point if the model's scoring logic is clinically justified and reasonable based on narrative details, rather than requiring an exact numeric match.
- Naranjo interpretation: Accept equivalent categories (Doubtful/Possible/Probable/Definite).

You must score the model's overall clinical accuracy on a scale of 1 to 5:
- 5 (Excellent): Clinically flawless. Both reasoning and JSON values are fully correct and match or are fully equivalent to the Ground Truth.
- 4 (Good): Clinically correct and logical, but not an exact match (e.g., minor synonym for MedDRA PT, or a slightly different Naranjo score that is still clinically reasonable and within +/- 1).
- 3 (Acceptable): Moderately sound but has minor errors in reasoning or Naranjo scoring that do not affect the main clinical conclusion (seriousness/expectedness).
- 2 (Poor): Significant clinical errors or flaws in Naranjo logic, but contains some correct elements.
- 1 (Unacceptable): Completely inaccurate, major clinical hallucinations, or fails basic safety mismatch rules.

Respond ONLY with a JSON block matching the requested schema."""


def build_judge_prompt(user_msg: str, gt_text: str, pred_text: str) -> str:
    return f"""Evaluate the Model Predicted Response (clinical reasoning + JSON) against the Ground Truth and Patient Narrative + RSI.

Context (Patient Narrative + RSI):
{user_msg}

Ground Truth Response:
{gt_text}

Model Response to Evaluate:
{pred_text}

Judge the Model's output and reasoning. Respond with ONLY a raw JSON object."""


# ── Main judge call with key-rotation (mirrors generate_reviews.py) ─
def call_judge(prompt: str) -> dict | None:
    """Calls the LLM judge with round-robin key rotation and failure tracking."""
    success = False
    retry_count = 0
    result = None

    while not success:
        key_val, client = client_queue.get()
        try:
            # Enforce minimum interval between requests per key
            last_used = key_last_used.get(key_val, 0.0)
            wait_time = MIN_INTERVAL - (time.time() - last_used)
            if wait_time > 0:
                time.sleep(wait_time)
            key_last_used[key_val] = time.time()

            resp = client.models.generate_content(
                model=MODEL_NAME,
                contents=prompt,
                config=types.GenerateContentConfig(
                    response_mime_type="application/json",
                    response_schema=LLMJudgeResponse,
                    system_instruction=JUDGE_SYSTEM,
                )
            )

            if resp.text:
                result = json.loads(resp.text.strip())
                success = True
                handle_success(key_val)
            else:
                raise Exception("Empty response returned from API.")

        except json.JSONDecodeError:
            handle_key_failure(key_val, "JSON decode error")
            retry_count += 1
            if retry_count >= 3:
                print("  Fatal: JSON decode error after 3 retries.")
                break
            time.sleep(retry_count * 2)

        except Exception as e:
            error_str = str(e)
            handle_key_failure(key_val, error_str)
            is_rate_limit = (
                "429" in error_str
                or "quota" in error_str.lower()
                or "resourceexhausted" in error_str.lower()
                or "exhausted" in error_str.lower()
                or "rate" in error_str.lower()
            )
            if is_rate_limit:
                handle_rate_limit(key_val)
            else:
                retry_count += 1
                if retry_count >= 3:
                    print(f"  Fatal Error: {error_str[:120]}. Skipping.")
                    break
                time.sleep(retry_count * 5)

        finally:
            if key_val not in disabled_keys:
                client_queue.put((key_val, client))

    return result


# ── Excel writer ───────────────────────────────────────────────────
def write_excel(results: list, total: int):
    wb = openpyxl.Workbook()

    # Sheet 1: Comparison Logs
    ws1 = wb.active
    ws1.title = "Comparison Logs"

    headers = [
        "Row #", "Suspected Drug", "Narrative (Snippet)",
        "GT is_serious", "Pred is_serious", "Seriousness Match",
        "GT Criteria", "Pred Criteria", "Criteria Match",
        "GT MedDRA PT", "Pred MedDRA PT", "MedDRA Match",
        "GT Expectedness", "Pred Expectedness", "Expectedness Match",
        "GT Naranjo Score", "Pred Naranjo Score", "Naranjo Score Match",
        "GT Interpretation", "Pred Interpretation", "Interpretation Match",
        "Clinical Accuracy Score", "Judge Clinical Explanation", "Judge Error",
    ]

    for ci, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = THIN_BORDER

    for ri, res in enumerate(results, start=2):
        vals = [
            res["row"], res["drug"], res["narrative"],
            res.get("gt_is_serious"), res.get("pr_is_serious"), res.get("serious_match"),
            res.get("gt_criteria"),  res.get("pr_criteria"),   res.get("criteria_match"),
            res.get("gt_meddra"),    res.get("pr_meddra"),     res.get("meddra_match"),
            res.get("gt_expect"),    res.get("pr_expect"),     res.get("expect_match"),
            res.get("gt_naranjo"),   res.get("pr_naranjo"),    res.get("naranjo_match"),
            res.get("gt_interp"),    res.get("pr_interp"),     res.get("interp_match"),
            res.get("clinical_accuracy_score"),
            res.get("clinical_explanation", ""),
            res.get("judge_error", False),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=ci, value=v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            # Color boolean match cells (skip raw value cols and score/explanation/error)
            if isinstance(v, bool) and ci not in [4, 5, 23, 24]:
                cell.fill = GREEN if v else RED
            # Color clinical accuracy score
            if ci == 22:
                try:
                    s = int(v)
                    cell.fill = GREEN if s >= 4 else (YELLOW if s == 3 else RED)
                except (ValueError, TypeError):
                    pass

    for ci in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(ci)].width = min(max(len(headers[ci-1]) + 4, 14), 40)
    ws1.column_dimensions["C"].width = 60
    ws1.column_dimensions["W"].width = 50
    ws1.auto_filter.ref = ws1.dimensions
    ws1.freeze_panes = "A2"

    # Sheet 2: Analytics Summary
    ws2 = wb.create_sheet("Analytics Summary")
    valid       = [r for r in results if not r.get("judge_error", False)]
    valid_count = len(valid)
    err_count   = len(results) - valid_count

    def acc(key):
        return sum(1 for r in valid if r.get(key)) / valid_count if valid_count else 0

    def cnt(key):
        return sum(1 for r in valid if r.get(key))

    scores         = [int(r["clinical_accuracy_score"]) for r in valid if r.get("clinical_accuracy_score") is not None]
    avg_score      = sum(scores) / len(scores) if scores else 0.0
    accuracy_count = sum(1 for s in scores if s >= 4)
    accuracy_rate  = accuracy_count / len(scores) if scores else 0.0
    score_counts   = Counter(scores)

    row = 1
    ws2.cell(row=row, column=1, value="PV MODEL EVALUATION – LLM JUDGE ANALYTICS").font = Font(bold=True, size=14)
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    row = 3
    for ci, h in enumerate(["Metric", "Count", "Percentage"], 1):
        c = ws2.cell(row=row, column=ci, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = THIN_BORDER

    data = [
        ("Total Samples",                   total,          ""),
        ("Successfully Judged",             valid_count,    f"{valid_count/total*100:.1f}%" if total else ""),
        ("Judge Errors / Skipped",          err_count,      f"{err_count/total*100:.1f}%"   if total else ""),
        ("", "", ""),
        ("── FIELD-LEVEL ACCURACY (Semantic) ──", "", ""),
        ("Seriousness (is_serious)",        cnt("serious_match"),  f"{acc('serious_match')*100:.1f}%"),
        ("Seriousness Criteria",            cnt("criteria_match"), f"{acc('criteria_match')*100:.1f}%"),
        ("MedDRA PT (Semantic)",            cnt("meddra_match"),   f"{acc('meddra_match')*100:.1f}%"),
        ("Expectedness",                    cnt("expect_match"),   f"{acc('expect_match')*100:.1f}%"),
        ("Naranjo Score (Semantic)",        cnt("naranjo_match"),  f"{acc('naranjo_match')*100:.1f}%"),
        ("Naranjo Interpretation",          cnt("interp_match"),   f"{acc('interp_match')*100:.1f}%"),
        ("", "", ""),
        ("── CLINICAL ACCURACY SCORES ──", "", ""),
        ("Average Clinical Accuracy Score", f"{avg_score:.2f} / 5.00", ""),
        ("Clinical Accuracy Rate (Score >= 4)", accuracy_count,  f"{accuracy_rate*100:.1f}%"),
        ("Score 5 (Excellent)",             score_counts[5], f"{score_counts[5]/len(scores)*100:.1f}%" if scores else ""),
        ("Score 4 (Good)",                  score_counts[4], f"{score_counts[4]/len(scores)*100:.1f}%" if scores else ""),
        ("Score 3 (Acceptable)",            score_counts[3], f"{score_counts[3]/len(scores)*100:.1f}%" if scores else ""),
        ("Score 2 (Poor)",                  score_counts[2], f"{score_counts[2]/len(scores)*100:.1f}%" if scores else ""),
        ("Score 1 (Unacceptable)",          score_counts[1], f"{score_counts[1]/len(scores)*100:.1f}%" if scores else ""),
    ]

    for ri, (metric, val, pct) in enumerate(data, start=row + 1):
        ws2.cell(row=ri, column=1, value=metric).font = BOLD if "──" in str(metric) else Font(size=11)
        ws2.cell(row=ri, column=2, value=val).border = THIN_BORDER
        ws2.cell(row=ri, column=3, value=pct).border = THIN_BORDER
        ws2.cell(row=ri, column=1).border = THIN_BORDER
        if isinstance(pct, str) and "%" in pct:
            pv = float(pct.replace("%", ""))
            ws2.cell(row=ri, column=3).fill = GREEN if pv >= 80 else (YELLOW if pv >= 60 else RED)

    # Confusion matrices
    section = row + len(data) + 3
    for title, gt_key, pr_key in [
        ("EXPECTEDNESS CONFUSION MATRIX",             "gt_expect", "pr_expect"),
        ("CAUSALITY INTERPRETATION CONFUSION MATRIX", "gt_interp", "pr_interp"),
    ]:
        ws2.cell(row=section, column=1, value=title).font = Font(bold=True, size=12)
        norm   = lambda v: str(v).strip().lower()
        labels = sorted(set(norm(r.get(gt_key, "")) for r in valid) | set(norm(r.get(pr_key, "")) for r in valid))
        c = ws2.cell(row=section+1, column=1, value="GT \\ Pred")
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = THIN_BORDER
        for ci, lbl in enumerate(labels, 2):
            c = ws2.cell(row=section+1, column=ci, value=lbl)
            c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = THIN_BORDER
        for ri2, gl in enumerate(labels, start=section+2):
            ws2.cell(row=ri2, column=1, value=gl).font = BOLD
            ws2.cell(row=ri2, column=1).border = THIN_BORDER
            for ci, pl in enumerate(labels, 2):
                n    = sum(1 for r in valid if norm(r.get(gt_key, "")) == gl and norm(r.get(pr_key, "")) == pl)
                cell = ws2.cell(row=ri2, column=ci, value=n)
                cell.border = THIN_BORDER
                if gl == pl:
                    cell.fill = GREEN
        section = section + len(labels) + 4

    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 18

    wb.save(OUTPUT)
    print(f"\n  Results saved to: {OUTPUT}")


# ── Main ───────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--keys", type=str, default="", help="Comma-separated Gemini API keys (overrides .env)")
    args = parser.parse_args()

    # Collect keys: CLI arg > .env / env var
    if args.keys:
        keys = [k.strip() for k in args.keys.split(",") if k.strip()]
    else:
        keys = load_api_keys()

    if not keys:
        print("ERROR: No API keys found. Set GEMINI_API_KEYS in .env or pass --keys.")
        sys.exit(1)

    init_key_manager(keys)

    # Load dataset
    rows = []
    with open(DATASET, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    print(f"Loaded {len(rows)} samples.")

    # Load checkpoint
    completed: dict = {}
    if os.path.exists(CHECKPOINT):
        with open(CHECKPOINT, "r") as f:
            completed = json.load(f)
        print(f"Resuming from checkpoint ({len(completed)} already judged).")

    results = []
    for idx, row in enumerate(rows, start=1):
        msgs = row.get("messages", [])
        user_msg = gt_text = pred_text = ""
        for m in msgs:
            if m["role"] == "user":
                user_msg = m["content"]
            elif m["role"] == "assistant":
                gt_text = m["content"]
            elif m["role"] == "trained_model_response":
                pred_text = m["content"]

        # Extract narrative snippet and drug name
        narr_m = re.search(r"Patient Narrative:\s*(.+?)(?:\n\nReference Safety|$)", user_msg, re.DOTALL)
        narrative = narr_m.group(1)[:200].strip() if narr_m else user_msg[:200]
        drug_m = re.search(r"The suspected drug is ([^.]+)\.", user_msg)
        drug = drug_m.group(1).strip() if drug_m else "N/A"

        gt_json = extract_json_block(gt_text)
        pr_json = extract_json_block(pred_text)

        base = {"row": idx, "drug": drug, "narrative": narrative}

        if gt_json is None or pr_json is None:
            base["judge_error"] = True
            results.append(base)
            continue

        # Populate GT/Pred values
        base.update({
            "gt_is_serious": gt_json.get("seriousness", {}).get("is_serious"),
            "pr_is_serious": pr_json.get("seriousness", {}).get("is_serious"),
            "gt_criteria":   gt_json.get("seriousness", {}).get("criteria", ""),
            "pr_criteria":   pr_json.get("seriousness", {}).get("criteria", ""),
            "gt_meddra":     gt_json.get("meddra_pt", ""),
            "pr_meddra":     pr_json.get("meddra_pt", ""),
            "gt_expect":     gt_json.get("expectedness", ""),
            "pr_expect":     pr_json.get("expectedness", ""),
            "gt_naranjo":    gt_json.get("causality", {}).get("naranjo_score"),
            "pr_naranjo":    pr_json.get("causality", {}).get("naranjo_score"),
            "gt_interp":     gt_json.get("causality", {}).get("interpretation", ""),
            "pr_interp":     pr_json.get("causality", {}).get("interpretation", ""),
        })

        # Load from checkpoint if already judged
        str_idx = str(idx)
        if str_idx in completed:
            j = completed[str_idx]
            base.update({
                "serious_match":          j.get("seriousness_is_serious_match"),
                "criteria_match":         j.get("seriousness_criteria_match"),
                "meddra_match":           j.get("meddra_pt_match"),
                "expect_match":           j.get("expectedness_match"),
                "naranjo_match":          j.get("naranjo_score_match"),
                "interp_match":           j.get("naranjo_interpretation_match"),
                "clinical_explanation":   j.get("clinical_explanation", ""),
                "clinical_accuracy_score":j.get("clinical_accuracy_score"),
                "judge_error":            False,
            })
            results.append(base)
            continue

        # Call the LLM judge
        prompt = build_judge_prompt(user_msg, gt_text, pred_text)
        print(f"  [{idx:3d}/{len(rows)}] Judging... (drug: {drug[:40]})", end="", flush=True)

        judge_result = call_judge(prompt)

        if judge_result is None:
            print(" FAIL (judge failed)")
            base["judge_error"] = True
            results.append(base)
            continue

        base.update({
            "serious_match":           judge_result.get("seriousness_is_serious_match"),
            "criteria_match":          judge_result.get("seriousness_criteria_match"),
            "meddra_match":            judge_result.get("meddra_pt_match"),
            "expect_match":            judge_result.get("expectedness_match"),
            "naranjo_match":           judge_result.get("naranjo_score_match"),
            "interp_match":            judge_result.get("naranjo_interpretation_match"),
            "clinical_explanation":    judge_result.get("clinical_explanation", ""),
            "clinical_accuracy_score": judge_result.get("clinical_accuracy_score"),
            "judge_error":             False,
        })
        results.append(base)

        # Save checkpoint
        completed[str_idx] = judge_result
        with open(CHECKPOINT, "w") as f:
            json.dump(completed, f)

        print(f" OK (score={judge_result.get('clinical_accuracy_score')})")

    # Write Excel
    print(f"\nWriting Excel with {len(results)} results...")
    write_excel(results, len(rows))

    # Print final summary
    valid = [r for r in results if not r.get("judge_error", False)]
    vc    = len(valid)
    if vc:
        def a(k): return sum(1 for r in valid if r.get(k)) / vc * 100
        scores         = [int(r["clinical_accuracy_score"]) for r in valid if r.get("clinical_accuracy_score") is not None]
        avg_score      = sum(scores) / len(scores) if scores else 0.0
        accuracy_count = sum(1 for s in scores if s >= 4)
        accuracy_rate  = accuracy_count / len(scores) if scores else 0.0

        print(f"\n{'='*60}")
        print(f"  LLM JUDGE EVALUATION COMPLETE")
        print(f"  Judged: {vc}/{len(rows)}")
        print(f"{'='*60}")
        print(f"  FIELD ACCURACIES (Semantic):")
        print(f"    Seriousness:       {a('serious_match'):.1f}%")
        print(f"    Criteria:          {a('criteria_match'):.1f}%")
        print(f"    MedDRA PT:         {a('meddra_match'):.1f}%")
        print(f"    Expectedness:      {a('expect_match'):.1f}%")
        print(f"    Naranjo Score:     {a('naranjo_match'):.1f}%")
        print(f"    Naranjo Interp:    {a('interp_match'):.1f}%")
        print(f"  CLINICAL ACCURACY METRICS:")
        print(f"    Average Score:     {avg_score:.2f} / 5.00")
        print(f"    Accuracy Rate (>=4): {accuracy_rate*100:.1f}%")
        print(f"{'='*60}")


if __name__ == "__main__":
    main()
