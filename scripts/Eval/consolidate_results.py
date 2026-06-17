"""
Consolidated Accuracy Report
=============================
Reads both evaluation Excel files (manual + LLM judge), merges them,
rates each sample as Accurate/Inaccurate per method, and produces a
final consolidated Excel with:
  Sheet 1 – Per-Sample Ratings (every row rated by both methods)
  Sheet 2 – Overall Accuracy Summary & Comparison
"""

import os
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────
BASE = os.path.dirname(__file__)
MANUAL_XL = os.path.join(BASE, "pv_evaluation_results.xlsx")
LLM_XL    = os.path.join(BASE, "pv_evaluation_llm_judge.xlsx")
OUTPUT    = os.path.join(BASE, "pv_consolidated_accuracy.xlsx")

# ── Styles ─────────────────────────────────────────────────────────
GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BLUE_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
DARK_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD = Font(bold=True, size=11)
BOLD_14 = Font(bold=True, size=14)
BOLD_12 = Font(bold=True, size=12)
NORMAL = Font(size=11)
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def read_comparison_sheet(path):
    """Read the 'Comparison Logs' sheet and return rows as list of dicts."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Comparison Logs"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    data = []
    for row in rows[1:]:
        d = {}
        for i, val in enumerate(row):
            if i < len(headers):
                d[headers[i]] = val
        data.append(d)
    wb.close()
    return data


def safe_bool(val):
    """Convert various truthy representations to bool."""
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip().lower() == "true"
    return False


def main():
    print("Reading manual evaluation...")
    manual = read_comparison_sheet(MANUAL_XL)
    print(f"  -> {len(manual)} rows")

    print("Reading LLM judge evaluation...")
    llm = read_comparison_sheet(LLM_XL)
    print(f"  -> {len(llm)} rows")

    # Build lookup by row number
    llm_by_row = {r.get("Row #"): r for r in llm}

    # ── Define which columns hold match booleans ───────────────
    # Manual sheet columns
    manual_fields = {
        "Seriousness":    "Seriousness Match",
        "Criteria":       "Criteria Match",
        "MedDRA PT":      "MedDRA Match",
        "Expectedness":   "Expectedness Match",
        "Naranjo Score":  "Naranjo Score Match",
        "Interpretation": "Interpretation Match",
    }
    # LLM sheet columns (same names)
    llm_fields = manual_fields.copy()

    # ── Build consolidated data ────────────────────────────────
    consolidated = []
    for m in manual:
        row_num = m.get("Row #")
        l = llm_by_row.get(row_num, {})

        entry = {
            "row": row_num,
            "drug": m.get("Suspected Drug", ""),
            "narrative": m.get("Narrative (Snippet)", ""),
            # GT values
            "gt_serious": m.get("GT is_serious"),
            "gt_criteria": m.get("GT Criteria", ""),
            "gt_meddra": m.get("GT MedDRA PT", ""),
            "gt_expect": m.get("GT Expectedness", ""),
            "gt_naranjo": m.get("GT Naranjo Score"),
            "gt_interp": m.get("GT Interpretation", ""),
            # Model values
            "pr_serious": m.get("Pred is_serious"),
            "pr_criteria": m.get("Pred Criteria", ""),
            "pr_meddra": m.get("Pred MedDRA PT", ""),
            "pr_expect": m.get("Pred Expectedness", ""),
            "pr_naranjo": m.get("Pred Naranjo Score"),
            "pr_interp": m.get("Pred Interpretation", ""),
        }

        # Manual matches
        m_matches = {}
        for label, col in manual_fields.items():
            m_matches[label] = safe_bool(m.get(col))
        entry["manual_matches"] = m_matches
        m_total = sum(m_matches.values())
        entry["manual_score"] = f"{m_total}/{len(m_matches)}"
        entry["manual_accurate"] = (m_total == len(m_matches))

        # LLM matches & scores
        l_matches = {}
        for label, col in llm_fields.items():
            l_matches[label] = safe_bool(l.get(col))
        entry["llm_matches"] = l_matches
        
        # Clinical Accuracy Score (1-5)
        try:
            score_val = int(l.get("Clinical Accuracy Score"))
        except (ValueError, TypeError):
            score_val = None
        entry["clinical_accuracy_score"] = score_val
        
        # Accurate if score >= 4 (5 is perfect, 4 is clinically correct but not exact match)
        entry["llm_accurate"] = (score_val >= 4) if score_val is not None else False
        entry["llm_score"] = f"{score_val}/5" if score_val is not None else "N/A"
        
        entry["llm_error"] = safe_bool(l.get("Judge Error"))
        entry["clinical_explanation"] = l.get("Judge Clinical Explanation", "")

        consolidated.append(entry)

    # ── Write Excel ────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════
    # Sheet 1: Per-Sample Ratings
    # ══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Per-Sample Ratings"

    headers = [
        "Row #", "Suspected Drug", "Narrative (Snippet)",
        # GT
        "GT Serious", "GT Criteria", "GT MedDRA PT",
        "GT Expectedness", "GT Naranjo", "GT Interpretation",
        # Model
        "Model Serious", "Model Criteria", "Model MedDRA PT",
        "Model Expectedness", "Model Naranjo", "Model Interpretation",
        # Manual eval
        "M: Seriousness", "M: Criteria", "M: MedDRA",
        "M: Expectedness", "M: Naranjo", "M: Interp",
        "Manual Score", "Manual Rating",
        # LLM eval
        "L: Seriousness", "L: Criteria", "L: MedDRA",
        "L: Expectedness", "L: Naranjo", "L: Interp",
        "LLM Score (1-5)", "LLM Rating",
        # Extra
        "Judge Clinical Explanation", "Agreement",
    ]

    for ci, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.fill = BLUE_FILL
        c.font = WHITE_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = THIN

    for ri, e in enumerate(consolidated, start=2):
        mm = e["manual_matches"]
        lm = e["llm_matches"]

        vals = [
            e["row"], e["drug"], e["narrative"],
            e["gt_serious"], e["gt_criteria"], e["gt_meddra"],
            e["gt_expect"], e["gt_naranjo"], e["gt_interp"],
            e["pr_serious"], e["pr_criteria"], e["pr_meddra"],
            e["pr_expect"], e["pr_naranjo"], e["pr_interp"],
            # Manual matches
            mm["Seriousness"], mm["Criteria"], mm["MedDRA PT"],
            mm["Expectedness"], mm["Naranjo Score"], mm["Interpretation"],
            e["manual_score"],
            "ACCURATE" if e["manual_accurate"] else "INACCURATE",
            # LLM matches
            lm["Seriousness"], lm["Criteria"], lm["MedDRA PT"],
            lm["Expectedness"], lm["Naranjo Score"], lm["Interpretation"],
            e["llm_score"],
            "ACCURATE" if e["llm_accurate"] else ("ERROR" if e["llm_error"] else "INACCURATE"),
            e["clinical_explanation"],
            "AGREE" if e["manual_accurate"] == e["llm_accurate"] else "DISAGREE",
        ]

        for ci, v in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=ci, value=v)
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Color boolean match cells
            if isinstance(v, bool) and 16 <= ci <= 21:  # Manual match cols
                cell.fill = GREEN if v else RED
            if isinstance(v, bool) and 24 <= ci <= 29:  # LLM match cols
                cell.fill = GREEN if v else RED
            # Color rating cells
            if ci == 23:  # Manual Rating
                cell.fill = GREEN if v == "ACCURATE" else RED
                cell.font = BOLD
            if ci == 31:  # LLM Rating
                cell.fill = GREEN if v == "ACCURATE" else (YELLOW if v == "ERROR" else RED)
                cell.font = BOLD
            if ci == 33:  # Agreement
                cell.fill = GREEN if v == "AGREE" else YELLOW
                cell.font = BOLD

    # Column widths
    widths = {1: 8, 2: 25, 3: 50, 23: 14, 30: 16, 31: 14, 32: 40, 33: 12}
    for ci in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(ci)].width = widths.get(ci, 14)
    ws1.auto_filter.ref = ws1.dimensions
    ws1.freeze_panes = "A2"

    # ══════════════════════════════════════════════════════════
    # Sheet 2: Overall Accuracy Summary
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Accuracy Summary")
    total = len(consolidated)
    valid_llm = [e for e in consolidated if not e["llm_error"]]

    def pct(n, d):
        return f"{n/d*100:.1f}%" if d else "N/A"

    row = 1
    ws2.cell(row=row, column=1, value="CONSOLIDATED ACCURACY REPORT").font = BOLD_14
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

    # ── Overview ───────────────────────────────────────────────
    row = 3
    ws2.cell(row=row, column=1, value="OVERVIEW").font = BOLD_12
    row = 4
    for ci, h in enumerate(["Metric", "Value"], 1):
        c = ws2.cell(row=row, column=ci, value=h)
        c.fill = BLUE_FILL; c.font = WHITE_FONT; c.border = THIN

    overview = [
        ("Total Samples", total),
        ("LLM Judge Errors", total - len(valid_llm)),
    ]
    for ri, (m, v) in enumerate(overview, start=row + 1):
        ws2.cell(row=ri, column=1, value=m).border = THIN
        ws2.cell(row=ri, column=2, value=v).border = THIN

    # ── Accuracy Comparison Table ──────────────────────────────
    row = row + len(overview) + 3
    ws2.cell(row=row, column=1, value="ACCURACY COMPARISON: MANUAL vs LLM JUDGE").font = BOLD_12
    row += 1
    comp_headers = ["Field", "Manual Eval – Correct", "Manual Eval – Accuracy",
                    "LLM Judge – Correct", "LLM Judge – Accuracy"]
    for ci, h in enumerate(comp_headers, 1):
        c = ws2.cell(row=row, column=ci, value=h)
        c.fill = DARK_FILL; c.font = WHITE_FONT; c.border = THIN
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    field_keys = ["Seriousness", "Criteria", "MedDRA PT", "Expectedness",
                  "Naranjo Score", "Interpretation"]
    for ri, fk in enumerate(field_keys, start=row + 1):
        m_correct = sum(1 for e in consolidated if e["manual_matches"].get(fk))
        l_correct = sum(1 for e in valid_llm if e["llm_matches"].get(fk))
        m_pct = pct(m_correct, total)
        l_pct = pct(l_correct, len(valid_llm))

        ws2.cell(row=ri, column=1, value=fk).font = BOLD
        ws2.cell(row=ri, column=1).border = THIN
        ws2.cell(row=ri, column=2, value=m_correct).border = THIN
        c3 = ws2.cell(row=ri, column=3, value=m_pct); c3.border = THIN
        ws2.cell(row=ri, column=4, value=l_correct).border = THIN
        c5 = ws2.cell(row=ri, column=5, value=l_pct); c5.border = THIN

        for c in [c3, c5]:
            v = float(c.value.replace("%", "")) if "%" in str(c.value) else 0
            c.fill = GREEN if v >= 80 else (YELLOW if v >= 60 else RED)

    # Overall row
    ri += 1
    m_all = sum(1 for e in consolidated if e["manual_accurate"])
    l_all = sum(1 for e in valid_llm if e["llm_accurate"])
    ws2.cell(row=ri, column=1, value="ALL FIELDS MATCH / CLINICALLY CORRECT").font = Font(bold=True, size=12)
    ws2.cell(row=ri, column=1).fill = LIGHT_BLUE; ws2.cell(row=ri, column=1).border = THIN
    ws2.cell(row=ri, column=2, value=m_all).fill = LIGHT_BLUE; ws2.cell(row=ri, column=2).border = THIN
    c3 = ws2.cell(row=ri, column=3, value=pct(m_all, total))
    c3.fill = LIGHT_BLUE; c3.border = THIN; c3.font = BOLD
    ws2.cell(row=ri, column=4, value=l_all).fill = LIGHT_BLUE; ws2.cell(row=ri, column=4).border = THIN
    c5 = ws2.cell(row=ri, column=5, value=pct(l_all, len(valid_llm)))
    c5.fill = LIGHT_BLUE; c5.border = THIN; c5.font = BOLD

    # ── Agreement Analysis ─────────────────────────────────────
    row = ri + 3
    ws2.cell(row=row, column=1, value="INTER-METHOD AGREEMENT").font = BOLD_12
    row += 1
    agree_headers = ["Category", "Count", "Percentage"]
    for ci, h in enumerate(agree_headers, 1):
        c = ws2.cell(row=row, column=ci, value=h)
        c.fill = BLUE_FILL; c.font = WHITE_FONT; c.border = THIN

    both_acc  = sum(1 for e in valid_llm if e["manual_accurate"] and e["llm_accurate"])
    both_inacc = sum(1 for e in valid_llm if not e["manual_accurate"] and not e["llm_accurate"])
    m_only = sum(1 for e in valid_llm if e["manual_accurate"] and not e["llm_accurate"])
    l_only = sum(1 for e in valid_llm if not e["manual_accurate"] and e["llm_accurate"])
    vl = len(valid_llm)

    agree_data = [
        ("Both rate ACCURATE", both_acc, pct(both_acc, vl)),
        ("Both rate INACCURATE", both_inacc, pct(both_inacc, vl)),
        ("Manual=ACCURATE, LLM=INACCURATE", m_only, pct(m_only, vl)),
        ("Manual=INACCURATE, LLM=ACCURATE", l_only, pct(l_only, vl)),
        ("", "", ""),
        ("Total Agreement", both_acc + both_inacc, pct(both_acc + both_inacc, vl)),
    ]
    for ri2, (lbl, cnt, p) in enumerate(agree_data, start=row + 1):
        ws2.cell(row=ri2, column=1, value=lbl).border = THIN
        ws2.cell(row=ri2, column=2, value=cnt).border = THIN
        ws2.cell(row=ri2, column=3, value=p).border = THIN
        if lbl == "Total Agreement":
            for c in range(1, 4):
                ws2.cell(row=ri2, column=c).font = BOLD
                ws2.cell(row=ri2, column=c).fill = LIGHT_BLUE

    # ── Score Distribution ─────────────────────────────────────
    row = ri2 + 3
    ws2.cell(row=row, column=1, value="LLM JUDGE CLINICAL SCORE DISTRIBUTION").font = BOLD_12
    row += 1
    for ci, h in enumerate(["Clinical Category", "Clinical Score", "Count", "Percentage"], 1):
        c = ws2.cell(row=row, column=ci, value=h)
        c.fill = BLUE_FILL; c.font = WHITE_FONT; c.border = THIN
        
    scores = [e["clinical_accuracy_score"] for e in valid_llm if e["clinical_accuracy_score"] is not None]
    score_counts = Counter(scores)
    dist_rows = [
        ("Excellent (Perfect GT equivalence)", 5, score_counts[5], pct(score_counts[5], len(scores))),
        ("Good (Clinically correct / minor diffs)", 4, score_counts[4], pct(score_counts[4], len(scores))),
        ("Acceptable (Minor clinical reasoning/scoring errors)", 3, score_counts[3], pct(score_counts[3], len(scores))),
        ("Poor (Significant clinical/Naranjo errors)", 2, score_counts[2], pct(score_counts[2], len(scores))),
        ("Unacceptable (Completely wrong / safety failure)", 1, score_counts[1], pct(score_counts[1], len(scores))),
    ]
    for ri4, (lbl, sc, count, p) in enumerate(dist_rows, start=row + 1):
        ws2.cell(row=ri4, column=1, value=lbl).border = THIN
        ws2.cell(row=ri4, column=2, value=sc).border = THIN
        ws2.cell(row=ri4, column=3, value=count).border = THIN
        ws2.cell(row=ri4, column=4, value=p).border = THIN

    # ── Final Accuracy Score ───────────────────────────────────
    row = ri4 + 3
    ws2.cell(row=row, column=1, value="FINAL ACCURACY SCORES").font = Font(bold=True, size=14, color="2F5496")
    row += 1
    
    avg_score = sum(scores) / len(scores) if scores else 0.0
    final = [
        ("Manual Evaluation Accuracy (all fields exact match)", pct(m_all, total)),
        ("LLM Judge Clinical Accuracy Rate (Score >= 4)", pct(l_all, len(valid_llm))),
        ("Average Clinical Accuracy Score (out of 5.0)", f"{avg_score:.2f} / 5.0"),
        ("Inter-Method Agreement Rate", pct(both_acc + both_inacc, vl)),
    ]
    for ri3, (lbl, val) in enumerate(final, start=row):
        ws2.cell(row=ri3, column=1, value=lbl).font = BOLD
        ws2.cell(row=ri3, column=1).border = THIN
        c = ws2.cell(row=ri3, column=2, value=val)
        c.font = Font(bold=True, size=14)
        c.border = THIN
        v = float(val.replace("%", "").split("/")[0].strip()) if "%" in val or "/" in val else 0
        if "%" in val:
            c.fill = GREEN if v >= 80 else (YELLOW if v >= 60 else RED)
        elif "/" in val:
            c.fill = GREEN if avg_score >= 4.0 else (YELLOW if avg_score >= 3.0 else RED)

    # Column widths
    ws2.column_dimensions["A"].width = 48
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 22
    ws2.column_dimensions["D"].width = 22
    ws2.column_dimensions["E"].width = 22

    wb.save(OUTPUT)
    print(f"\n{'='*60}")
    print(f"  CONSOLIDATED REPORT COMPLETE")
    print(f"  Samples: {total}")
    print(f"  Manual Accuracy (all match):  {pct(m_all, total)}")
    print(f"  LLM Judge Accuracy (Score>=4): {pct(l_all, len(valid_llm))}")
    print(f"  Average Score:                {avg_score:.2f}/5.0")
    print(f"  Agreement Rate:               {pct(both_acc + both_inacc, vl)}")
    print(f"{'='*60}")
    print(f"  Saved to: {OUTPUT}")


if __name__ == "__main__":
    main()
